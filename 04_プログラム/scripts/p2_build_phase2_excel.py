"""Phase 2評価CSVをExcel成果物へ統合する。"""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
PROGRAM_DIR = SCRIPT_DIR.parent
EVALUATION_DIR = PROGRAM_DIR / "output" / "sumo" / "evaluation"
OUTPUT_XLSX = EVALUATION_DIR / "phase2_results_excel.xlsx"

SOURCES = [
    ("Evacuation", "evacuation_summary.csv", "常総市シナリオA避難結果"),
    ("TrialSettings", "trial_settings_comparison.csv", "small / 10pct / full 試行設定比較"),
    ("CongestionLog", "congestion_log.csv", "60秒間隔の混雑ログ"),
    ("MajorRoutes", "major_route_congestion_summary.csv", "主要避難路別混雑集計"),
    ("P1P2_Joso", "phase1_phase2_comparison.csv", "常総市 Phase 1 / Phase 2 比較"),
    ("Municipalities", "evacuation_summary_by_municipality.csv", "市区町村別避難結果"),
    ("P1P2_Regions", "phase1_phase2_region_comparison.csv", "全域 Phase 1 / Phase 2 比較"),
]


def style_sheet(writer: pd.ExcelWriter, sheet_name: str) -> None:
    ws = writer.book[sheet_name]
    ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        width = min(max(max_len + 2, 10), 34)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def build_excel() -> None:
    EVALUATION_DIR.mkdir(parents=True, exist_ok=True)
    summary_rows = []
    frames: list[tuple[str, pd.DataFrame]] = []
    for sheet_name, file_name, label in SOURCES:
        path = EVALUATION_DIR / file_name
        df = pd.read_csv(path)
        frames.append((sheet_name, df))
        summary_rows.append(
            {
                "sheet": sheet_name,
                "source_file": file_name,
                "rows": len(df),
                "description": label,
            }
        )

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        summary = pd.DataFrame(summary_rows)
        summary.to_excel(writer, index=False, sheet_name="Summary")
        for sheet_name, df in frames:
            df.to_excel(writer, index=False, sheet_name=sheet_name)

        for sheet_name in ["Summary", *[sheet for sheet, _ in frames]]:
            style_sheet(writer, sheet_name)

        ws = writer.book["Summary"]
        ws.insert_rows(1, 2)
        ws["A1"] = "Phase 2 Excel成果物"
        ws["A2"] = "Phase 2評価CSVをExcelで確認しやすいように統合した固定成果物"
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
        ws["A1"].fill = PatternFill("solid", fgColor="0F766E")
        ws["A2"].font = Font(color="374151")

    print(f"[INFO] saved: {OUTPUT_XLSX}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("command", choices=["all"], help="task to run")
    args = parser.parse_args()
    if args.command == "all":
        build_excel()


if __name__ == "__main__":
    main()
